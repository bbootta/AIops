"""업무보고서 .xlsx 산출.

시트 구성 — 표지 / 목차 / 서식 14장 / 검증 / 산출근거.
값만 있는 엑셀은 감독당국 질의에 답하지 못한다. 라인마다 산식·규정근거·
산출 모듈을 같은 행에 두어, 시트 하나로 "이 숫자가 어디서 나왔는지"가 닫힌다.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from risk_lib.regulatory.forms import BuiltForm, submission_digest

_HEAD_FILL = PatternFill("solid", fgColor="1F3864")
_SUB_FILL = PatternFill("solid", fgColor="DDEBF7")
_FAIL_FILL = PatternFill("solid", fgColor="FCE4E4")
_TITLE_FONT = Font(name="맑은 고딕", size=14, bold=True, color="1F3864")
_HEAD_FONT = Font(name="맑은 고딕", size=9, bold=True, color="FFFFFF")
_BODY_FONT = Font(name="맑은 고딕", size=9)
_SUB_FONT = Font(name="맑은 고딕", size=9, bold=True)
_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_KRW_FMT = "#,##0"
_RATIO_FMT = "0.0000%"
_COUNT_FMT = "#,##0"

_COLUMNS = ("라인코드", "항목명", "단위", "값", "산식", "규정 근거", "산출 모듈")
_WIDTHS = (10, 46, 8, 20, 44, 40, 46)


def _write_header(ws, row: int, columns=_COLUMNS, widths=_WIDTHS) -> int:
    for c, (name, w) in enumerate(zip(columns, widths), start=1):
        cell = ws.cell(row=row, column=c, value=name)
        cell.fill, cell.font, cell.border = _HEAD_FILL, _HEAD_FONT, _BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(c)].width = w
    return row + 1


def _sheet_title(ws, title: str, subtitle: str, span: int = 7) -> int:
    ws.cell(row=1, column=1, value=title).font = _TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    c = ws.cell(row=2, column=1, value=subtitle)
    c.font = Font(name="맑은 고딕", size=9, color="595959")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=span)
    return 4


def _cover(wb: Workbook, built: list[BuiltForm], asof: str, meta: dict) -> None:
    ws = wb.create_sheet("표지", 0)
    ws.sheet_view.showGridLines = False
    ws.cell(row=2, column=2, value="업무보고서 (금융감독원 배포 기준)").font = \
        Font(name="맑은 고딕", size=18, bold=True, color="1F3864")
    ws.cell(row=3, column=2,
            value="RYNTA 리스크관리 에이전트 하네스 · 결정론적 산출").font = \
        Font(name="맑은 고딕", size=10, color="595959")
    rows = [
        ("제출기관", meta.get("institution", "(기관명)")),
        ("기준일", asof),
        ("작성 부서", "리스크관리부"),
        ("작성자", meta.get("prepared_by", "리스크관리부 실무자")),
        ("검토자", meta.get("reviewed_by", "리스크관리부장")),
        ("승인자", meta.get("approved_by", "리스크담당임원(CRO)")),
        ("서식 수", f"{len(built)}장"),
        ("라인 수", f"{sum(len(b.lines) for b in built):,}행"),
        ("검증 항목", f"{sum(len(b.checks) for b in built)}건 "
                      f"(실패 {sum(b.n_failed for b in built)}건)"),
        ("산출 지문 (SHA-256)", submission_digest(built)),
        ("난수 시드", str(meta.get("seed", 42))),
        ("산출 엔진", "risk_lib.pipeline.run_pipeline"),
    ]
    r = 6
    for label, value in rows:
        lc = ws.cell(row=r, column=2, value=label)
        lc.font = _SUB_FONT
        lc.fill = _SUB_FILL
        lc.border = _BORDER
        vc = ws.cell(row=r, column=3, value=value)
        vc.font = _BODY_FONT
        vc.border = _BORDER
        r += 1
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 78

    r += 1
    ws.cell(row=r, column=2, value="유의사항").font = _TITLE_FONT
    r += 1
    for note in (
        "① 서식 식별자(BR-01 …)는 내부 코드다. 금융감독원 배포본 서식번호가 "
        "확정되면 서식 식별자 매핑표 한 장만 교체하면 되고 라인코드·산식·근거는 "
        "그대로 쓴다. 배포본을 받지 않은 상태에서 서식번호를 지어내지 않았다.",
        "② 모든 값은 합성 포트폴리오에서 산출한 것이며 실제 기관 수치가 아니다. "
        "재현 방법: run_pipeline(generate_portfolio(seed), seed=seed, asof=기준일).",
        "③ 자산건전성 분류는 연체일수 기준 대용 규칙이다. 감독규정 제27조는 "
        "채무상환능력 평가를 함께 요구하므로 실제 제출 시 여신심사 판정이 선행되어야 한다.",
        "④ 각 서식의 소계·비율은 「검증」 시트에서 자체 대사한다. 실패 항목이 "
        "하나라도 있으면 해당 서식의 제출 상태는 draft로 남는다.",
        "⑤ 약어: BIS(국제결제은행 기준 자기자본비율) · CET1(보통주자본) · "
        "AT1(기타기본자본) · RWA(위험가중자산) · SA(표준방법) · IRB(내부등급법) · "
        "EAD(부도시익스포저) · PD(부도율) · LGD(부도시손실률) · EL(기대손실) · "
        "LCR(유동성커버리지비율) · HQLA(고유동성자산) · NSFR(순안정자금조달비율) · "
        "ASF(가용안정자금) · RSF(필요안정자금) · IRRBB(은행계정 금리리스크) · "
        "EVE(경제적가치) · NII(순이자이익) · ECL(기대신용손실) · BI(사업지표) · "
        "BIC(사업지표요소) · ILM(내부손실승수) · ORC(운영리스크 소요자기자본) · "
        "CRO(리스크담당임원).",
    ):
        c = ws.cell(row=r, column=2, value=note)
        c.font = Font(name="맑은 고딕", size=9)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        ws.row_dimensions[r].height = 46
        r += 1


def _toc(wb: Workbook, built: list[BuiltForm]) -> None:
    ws = wb.create_sheet("목차", 1)
    ws.sheet_view.showGridLines = False
    row = _sheet_title(ws, "목차", "서식별 제출 주기와 근거 규정", span=6)
    row = _write_header(ws, row,
                        ("서식", "서식명", "제출주기", "라인 수", "검증 실패",
                         "근거 규정"),
                        (10, 40, 10, 10, 10, 70))
    for b in built:
        vals = (b.spec.form_id, b.spec.form_name, b.spec.frequency,
                len(b.lines), b.n_failed, b.spec.citation)
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.font = _BODY_FONT
            cell.border = _BORDER
            if c == 5 and b.n_failed:
                cell.fill = _FAIL_FILL
        row += 1


def _form_sheet(wb: Workbook, b: BuiltForm) -> None:
    # 시트명은 31자 제한 + 특수문자 불가 — 서식ID로 짧게 고정한다.
    ws = wb.create_sheet(f"{b.spec.form_id}")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"
    row = _sheet_title(ws, f"[{b.spec.form_id}] {b.spec.form_name}",
                       f"제출주기 {b.spec.frequency} · 근거 {b.spec.citation}")
    row = _write_header(ws, row)
    for ln in b.lines:
        indent = "    " * ln.level
        cells = [
            ln.line_code, f"{indent}{ln.line_name}", ln.unit,
            ln.value if ln.unit != "text" else ln.text_value,
            ln.formula or "", ln.citation or "", ln.source_module or "",
        ]
        for c, v in enumerate(cells, start=1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.font = _SUB_FONT if ln.is_subtotal else _BODY_FONT
            cell.border = _BORDER
            if ln.is_subtotal:
                cell.fill = _SUB_FILL
            if c == 4:
                cell.number_format = {"KRW": _KRW_FMT, "ratio": _RATIO_FMT,
                                      "count": _COUNT_FMT}.get(ln.unit, "@")
                cell.alignment = Alignment(horizontal="right")
            if c in (5, 6, 7):
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        row += 1
    if any(ln.unit == "text" for ln in b.lines):
        for ln in b.lines:
            if ln.unit == "text" and ln.text_value:
                ws.cell(row=row + 1, column=2, value=f"※ {ln.text_value}").font = \
                    Font(name="맑은 고딕", size=8, italic=True, color="595959")
                row += 1


def _checks_sheet(wb: Workbook, built: list[BuiltForm]) -> None:
    ws = wb.create_sheet("검증")
    ws.sheet_view.showGridLines = False
    row = _sheet_title(
        ws, "서식 자체 대사",
        "소계 = 구성요소 합, 비율 = 분자 ÷ 분모. 실패가 하나라도 있으면 제출 불가.",
        span=6)
    row = _write_header(ws, row,
                        ("서식", "검증 항목", "기대값", "실제값", "차이", "판정"),
                        (10, 46, 22, 22, 18, 10))
    for b in built:
        for c in b.checks:
            vals = (b.spec.form_id, c.check_name, c.expected, c.actual,
                    c.diff, c.status)
            for i, v in enumerate(vals, start=1):
                cell = ws.cell(row=row, column=i, value=v)
                cell.font = _BODY_FONT
                cell.border = _BORDER
                if i in (3, 4, 5):
                    cell.number_format = "#,##0.00########"
                    cell.alignment = Alignment(horizontal="right")
                if c.status == "FAIL":
                    cell.fill = _FAIL_FILL
            row += 1


def _lineage_sheet(wb: Workbook, built: list[BuiltForm], asof: str,
                   meta: dict) -> None:
    ws = wb.create_sheet("산출근거")
    ws.sheet_view.showGridLines = False
    row = _sheet_title(
        ws, "산출근거 · 재현 절차",
        "같은 시드·같은 기준일이면 같은 지문이 나온다 — 그것이 재현의 정의다.",
        span=4)
    row = _write_header(ws, row, ("구분", "내용", "값", "참조"),
                        (18, 46, 40, 46))
    entries = [
        ("재현", "포트폴리오 생성", f"generate_portfolio(seed={meta.get('seed', 42)})",
         "risk_lib.data_gen"),
        ("재현", "파이프라인 실행",
         f"run_pipeline(portfolio, seed={meta.get('seed', 42)}, asof='{asof}')",
         "risk_lib.pipeline"),
        ("재현", "정규 테이블 실체화", "materialize_all + materialize_detail",
         "risk_lib.datamodel"),
        ("재현", "서식 채움", "build_forms(result, portfolio, tables)",
         "risk_lib.regulatory.forms"),
        ("지문", "제출본 SHA-256", submission_digest(built),
         "risk_lib.regulatory.forms.submission_digest"),
        ("통제", "직무분리",
         f"작성 {meta.get('prepared_by', '-')} / 승인 {meta.get('approved_by', '-')}",
         "AIMS_POLICY §2-4"),
        ("통제", "AI 자동확정 금지",
         "에이전트는 제안·검증만 수행하며 제출본을 자동 확정하지 않는다",
         "ISO/IEC 42001 A.9.2 · EU AI Act 제14조"),
    ]
    for kind, name, value, ref in entries:
        for i, v in enumerate((kind, name, value, ref), start=1):
            cell = ws.cell(row=row, column=i, value=v)
            cell.font = _BODY_FONT
            cell.border = _BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        row += 1
    row += 1
    ws.cell(row=row, column=1, value="서식 라인별 산출 모듈").font = _TITLE_FONT
    row += 1
    row = _write_header(ws, row, ("서식", "라인코드", "항목명", "산출 모듈"),
                        (10, 12, 46, 62))
    for b in built:
        for ln in b.lines:
            if not ln.source_module:
                continue
            for i, v in enumerate((b.spec.form_id, ln.line_code, ln.line_name,
                                   ln.source_module), start=1):
                cell = ws.cell(row=row, column=i, value=v)
                cell.font = _BODY_FONT
                cell.border = _BORDER
            row += 1


def write_workbook(built: list[BuiltForm], path: str | Path, *, asof: str,
                   meta: dict | None = None) -> Path:
    """서식 전체를 하나의 .xlsx로 쓴다. 반환값은 실제 기록된 경로."""
    meta = dict(meta or {})
    wb = Workbook()
    wb.remove(wb.active)
    _cover(wb, built, asof, meta)
    _toc(wb, built)
    for b in built:
        _form_sheet(wb, b)
    _checks_sheet(wb, built)
    _lineage_sheet(wb, built, asof, meta)
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    wb.save(p)
    return p
