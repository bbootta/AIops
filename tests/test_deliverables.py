"""산출물 패키징·업무보고서 (R10).

핵심 명제: 포장한 뒤 **검증하지 않으면** '만들었다'는 것만 알 뿐 '올바른지'는
모른다. ZIP은 자체 매니페스트로 자가검증돼야 한다.
"""

from __future__ import annotations

import hashlib
import warnings
import zipfile
from pathlib import Path

import pandas as pd
import pytest

from risk_lib.deliverables import (
    export_tables, export_ddl, export_catalog_summary,
    write_manifest, make_zip, verify_zip, build_deliverables,
)
from risk_lib.html_report import INSTITUTION_UNSPECIFIED, institution_label
from risk_lib.work_report import (
    ROUNDS, build_context, render_markdown, render_html, write_work_report,
)


@pytest.fixture(scope="module")
def tables(result, portfolio):
    from risk_lib.datamodel.materialize import materialize_all
    return materialize_all(result, portfolio)


# ----- 내보내기 ---------------------------------------------------------------

def test_every_table_exports_to_csv(tmp_path, tables):
    files = export_tables(tables, tmp_path / "t")
    assert len(files) == len(tables)
    for f in files:
        assert f.stat().st_size > 0
        back = pd.read_csv(f)
        assert len(back) == len(tables[f.stem])


def test_csv_uses_bom_so_excel_reads_korean(tmp_path, tables):
    """utf-8-sig가 아니면 Excel에서 한글이 깨진다 — 실사용에서 즉시 드러나는 결함."""
    files = export_tables({"rdm_obligor": tables["rdm_obligor"]}, tmp_path / "t")
    assert files[0].read_bytes()[:3] == b"\xef\xbb\xbf"


def test_ddl_orders_referenced_tables_first(tmp_path):
    """FK 대상이 뒤에 오면 스크립트를 그대로 실행할 수 없다."""
    sql = export_ddl(tmp_path).read_text(encoding="utf-8")
    pos = {}
    for line in sql.splitlines():
        if line.startswith("CREATE TABLE "):
            pos[line.split()[2].rstrip("(")] = len(pos)
    from risk_lib.datamodel import catalog as cat
    for spec in cat.ALL_TABLES:
        for fk in spec.foreign_keys:
            assert pos[fk.ref_table] < pos[spec.name], (
                f"{spec.name}이 참조 대상 {fk.ref_table}보다 먼저 생성된다")


def test_ddl_covers_every_catalog_table(tmp_path):
    from risk_lib.datamodel import catalog as cat
    sql = export_ddl(tmp_path).read_text(encoding="utf-8")
    for spec in cat.ALL_TABLES:
        assert f"CREATE TABLE {spec.name}" in sql


# ----- 무결성 매니페스트 ------------------------------------------------------

def test_manifest_hashes_match_files(tmp_path):
    (tmp_path / "a.txt").write_text("hello", encoding="utf-8")
    (tmp_path / "sub").mkdir()
    (tmp_path / "sub" / "b.txt").write_text("world", encoding="utf-8")
    m = write_manifest(tmp_path)
    lines = [l for l in m.read_text(encoding="utf-8").splitlines()
             if l and not l.startswith("#")]
    assert len(lines) == 2
    for line in lines:
        sha, size, rel = line.split(None, 2)
        data = (tmp_path / rel).read_bytes()
        assert hashlib.sha256(data).hexdigest() == sha
        assert int(size) == len(data)


def test_manifest_excludes_itself(tmp_path):
    (tmp_path / "a.txt").write_text("x", encoding="utf-8")
    m = write_manifest(tmp_path)
    assert "MANIFEST.txt" not in m.read_text(encoding="utf-8").split("형식")[-1]


# ----- ZIP 자가검증 -----------------------------------------------------------

def test_zip_verifies_against_its_own_manifest(tmp_path):
    src = tmp_path / "pkg"; src.mkdir()
    (src / "a.txt").write_text("alpha", encoding="utf-8")
    (src / "b.csv").write_text("x,y\n1,2\n", encoding="utf-8")
    write_manifest(src)
    z = make_zip(src, tmp_path / "out.zip")
    r = verify_zip(z)
    assert r["ok"] and r["n_verified"] == 2 and not r["mismatched"]


def test_zip_verification_detects_tampering(tmp_path):
    """검증이 발동 가능해야 통제다 — 변조본이 통과하면 안 된다."""
    src = tmp_path / "pkg"; src.mkdir()
    (src / "a.txt").write_text("alpha", encoding="utf-8")
    write_manifest(src)
    z = make_zip(src, tmp_path / "out.zip")

    # ZIP을 다시 만들며 내용만 바꾼다 (매니페스트는 그대로)
    tampered = tmp_path / "tampered.zip"
    with zipfile.ZipFile(z) as src_z, zipfile.ZipFile(tampered, "w") as dst:
        for n in src_z.namelist():
            data = src_z.read(n)
            if n == "a.txt":
                data = b"TAMPERED"
            dst.writestr(n, data)
    r = verify_zip(tampered)
    assert not r["ok"]
    assert "a.txt" in r["mismatched"]


def test_zip_without_manifest_is_not_ok(tmp_path):
    src = tmp_path / "pkg"; src.mkdir()
    (src / "a.txt").write_text("x", encoding="utf-8")
    z = make_zip(src, tmp_path / "no_manifest.zip")
    r = verify_zip(z)
    assert not r["ok"] and "MANIFEST" in r["reason"]


# ----- 업무보고서 -------------------------------------------------------------

def test_ten_rounds_are_declared():
    assert len(ROUNDS) == 10
    assert [r.no for r in ROUNDS] == list(range(1, 11))
    for r in ROUNDS:
        assert r.title and r.scope and r.artifacts


def test_work_report_states_open_items_not_just_successes(result, portfolio, tables):
    """미결 사항을 빼면 보고서가 아니라 홍보물이 된다."""
    ctx = build_context(result, portfolio, tables)
    md = render_markdown(ctx)
    assert "미결 사항" in md
    assert "backlog" in md
    # 커버리지 수치가 실제 값과 일치
    assert str(ctx["coverage"].get("backlog", 0)) in md
    assert f"{ctx['in_scope_ratio']*100:.1f}%" in md
    # 한계도 명시
    assert "규제 제출용이 아니다" in md


def test_work_report_records_defects_found_per_round(result, portfolio, tables):
    """라운드별 발견 결함이 보고서에 남아야 개선 이력이 추적된다."""
    md = render_markdown(build_context(result, portfolio, tables))
    with_findings = [r for r in ROUNDS if r.findings != "—"]
    assert len(with_findings) >= 4
    for r in with_findings:
        assert r.findings.split("·")[0].strip()[:12] in md


def test_work_report_numbers_match_the_system(result, portfolio, tables):
    from risk_lib.page_registry import PAGES
    from risk_lib.datamodel import catalog as cat
    ctx = build_context(result, portfolio, tables)
    assert ctx["n_tables"] == len(cat.ALL_TABLES)
    assert ctx["n_pages"] == len(PAGES)
    assert ctx["n_materialized"] == len(tables)
    assert ctx["schema_violations"] == 0


def test_work_report_renders_html(tmp_path, result, portfolio, tables):
    w = write_work_report(result, portfolio, tables, tmp_path)
    html = Path(w["html"]).read_text(encoding="utf-8")
    assert html.startswith("<!doctype html>")
    assert "업무보고서" in html and "<table" in html


# ----- 전체 패키징 ------------------------------------------------------------

@pytest.mark.slow
def test_full_package_builds_and_self_verifies(tmp_path, result, portfolio):
    out = build_deliverables(result, portfolio, tmp_path / "deliv",
                             zip_name="pkg.zip")
    assert out["schema_violations"] == 0
    assert out["zip_verified"]["ok"], out["zip_verified"]
    assert out["zip_verified"]["n_verified"] >= 100
    z = Path(out["zip"])
    assert z.exists() and z.stat().st_size > 100_000
    with zipfile.ZipFile(z) as zf:
        names = zf.namelist()
    for expected in ("README.md", "MANIFEST.txt",
                     "01_datamodel/schema.sql",
                     "02_reports/executive.html",
                     "04_work_report/업무보고서.html"):
        assert expected in names, f"{expected} 누락"


# ----- 기관 귀속 --------------------------------------------------------------
#
# 배포 패키지의 기관 귀속은 실행이 말한 institution_code 에서 와야 한다. 넘기지
# 않으면 업무보고서 표지가 "(기관명)"으로 조용히 메워져, 받는 쪽은 기관 귀속이
# 빠졌다는 사실 자체를 알 수 없었다.

def _stub_result(**meta):
    class _R:
        pass
    r = _R()
    r.meta = dict(meta)
    return r


def test_institution_label_warns_when_the_run_declared_none():
    with pytest.warns(UserWarning, match="기관코드"):
        label = institution_label(_stub_result(asof="2026-06-11", seed=42))
    assert label == INSTITUTION_UNSPECIFIED
    # 대체 기관코드를 표지에 적으면 실행이 한 적 없는 귀속이 배포본에 남는다.
    assert "KR_BANK" not in label


def test_institution_label_uses_the_declared_institution():
    with warnings.catch_warnings():
        warnings.simplefilter("error")          # 경고가 나면 실패한다
        assert institution_label(
            _stub_result(institution_code="KR_BANK_01")) == "KR_BANK_01"


def _cover_institution(xlsx_path) -> str:
    from openpyxl import load_workbook
    ws = load_workbook(xlsx_path)["표지"]
    for row in ws.iter_rows(values_only=True):
        cells = [c for c in row if c is not None]
        if cells and str(cells[0]) == "제출기관":
            return str(cells[1])
    raise AssertionError("업무보고서 표지에 제출기관 행이 없다")


@pytest.mark.slow
def test_package_without_institution_warns_and_says_so(tmp_path, result,
                                                       portfolio):
    """기관코드 없이 부르면 경고가 나고, 그 사실이 산출물에 적힌다."""
    with pytest.warns(UserWarning, match="기관코드"):
        out = build_deliverables(result, portfolio, tmp_path / "d",
                                 zip_name="pkg.zip")
    assert out["institution"] == INSTITUTION_UNSPECIFIED
    assert _cover_institution(out["regulatory_xlsx"]) == INSTITUTION_UNSPECIFIED
    readme = (Path(out["root"]) / "README.md").read_text(encoding="utf-8")
    assert "기관 미지정" in readme
    assert "run_pipeline(institution_code=...)" in readme


@pytest.mark.slow
def test_package_labels_the_institution_the_run_declared(tmp_path, result,
                                                         portfolio):
    """실행이 기관을 말하면 표지·README가 그 기관으로 라벨되고 경고가 없다."""
    import copy
    declared = copy.copy(result)
    declared.meta = {**result.meta, "institution_code": "KR_BANK_01"}

    with warnings.catch_warnings(record=True) as caught:
        warnings.simplefilter("always")
        out = build_deliverables(declared, portfolio, tmp_path / "d",
                                 zip_name="pkg.zip")
    assert not [w for w in caught if "기관코드" in str(w.message)]
    assert out["institution"] == "KR_BANK_01"
    assert out["institution_source"] == "실행 지정"
    assert _cover_institution(out["regulatory_xlsx"]) == "KR_BANK_01"
    readme = (Path(out["root"]) / "README.md").read_text(encoding="utf-8")
    assert "제출기관 **KR_BANK_01**" in readme
    assert "기관 미지정" not in readme
