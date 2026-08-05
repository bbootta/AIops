"""금감원 업무보고서 — 서식 구조·값·자체대사·엑셀 산출.

핵심 명제 세 가지:
  1) 서식 값은 파이프라인 공표값과 **정확히** 일치한다(따로 계산하지 않는다).
  2) 모든 소계·비율은 서식이 스스로 대사하고, 실패가 있으면 제출 상태로
     올라가지 않는다.
  3) 모든 라인은 산출 근거(모듈)를 갖거나, 근거가 없는 이유가 분명하다.
"""

from __future__ import annotations

import pytest
from openpyxl import load_workbook

from risk_lib.datamodel import catalog as cat
from risk_lib.datamodel.spec import validate
from risk_lib.regulatory import build_forms, form_frames, write_workbook
from risk_lib.regulatory.forms import FORMS, FormCheck, submission_digest


@pytest.fixture(scope="module")
def tables(result, portfolio):
    from risk_lib.ui_studio.studio import build_studio
    return build_studio(result, portfolio).tables


@pytest.fixture(scope="module")
def built(result, portfolio, tables):
    return build_forms(result, portfolio, tables)


# ----- 구조 -------------------------------------------------------------------

def test_every_form_builds_with_lines(built):
    assert len(built) == len(FORMS)
    for b in built:
        assert b.lines, f"{b.spec.form_id} 라인 없음"
        assert b.spec.citation, f"{b.spec.form_id} 근거 규정 없음"


def test_line_codes_unique_within_form(built):
    for b in built:
        codes = [ln.line_code for ln in b.lines]
        assert len(codes) == len(set(codes)), f"{b.spec.form_id} 라인코드 중복"


def test_sheet_order_is_unique_and_dense():
    orders = sorted(f.sheet_order for f in FORMS)
    assert orders == list(range(1, len(FORMS) + 1))


def test_units_are_declared_and_values_match_unit(built):
    for b in built:
        for ln in b.lines:
            assert ln.unit in ("KRW", "ratio", "multiple", "count", "text")
            if ln.unit == "text":
                assert ln.value is None and ln.text_value
            else:
                assert ln.value is not None, f"{b.spec.form_id}/{ln.line_code}"


def test_ratio_lines_are_fractions_not_percents(built):
    """비율은 0.115로 담는다. 11.5로 담으면 엑셀 서식이 1150%로 표시된다.

    회전율처럼 1을 크게 넘는 것이 정상인 값은 `multiple` 단위를 쓴다 — 여기서
    한계를 느슨하게 풀면 정작 잡아야 할 백분율 오기를 놓친다.
    """
    for b in built:
        for ln in b.lines:
            if ln.unit == "ratio" and ln.value is not None:
                assert -5.0 <= ln.value <= 5.0, f"{b.spec.form_id}/{ln.line_code}"


# ----- 자체 대사 --------------------------------------------------------------

def test_all_form_checks_pass(built):
    failed = [(b.spec.form_id, c.check_name, c.expected, c.actual)
              for b in built for c in b.checks if c.status == "FAIL"]
    assert failed == [], failed


def test_every_form_has_at_least_one_check(built):
    for b in built:
        assert b.checks, f"{b.spec.form_id} 자체대사 없음 — 제출 전 오류를 못 잡는다"


def test_failed_check_blocks_approval_status(built):
    """검증 실패가 있으면 제출 상태가 approved로 올라가지 않는다."""
    frames = form_frames(built, "2026-06-11")
    subs = frames["reg_submission"]
    assert (subs.loc[subs["n_failed_checks"] > 0, "status"] == "draft").all()
    assert (subs.loc[subs["n_failed_checks"] == 0, "status"] == "approved").all()


def test_check_status_is_tolerance_driven():
    assert FormCheck("t", 100.0, 100.5, tolerance=1.0).status == "PASS"
    assert FormCheck("t", 100.0, 102.0, tolerance=1.0).status == "FAIL"


# ----- 공표값 대사 ------------------------------------------------------------

def _line(built, form_id, code):
    b = next(x for x in built if x.spec.form_id == form_id)
    return next(ln for ln in b.lines if ln.line_code == code)


def test_capital_ratio_lines_match_pipeline(built, result):
    assert _line(built, "BR-01", "3100").value == pytest.approx(
        result.bis.cet1_ratio, rel=1e-15)
    assert _line(built, "BR-01", "3300").value == pytest.approx(
        result.bis.total_ratio, rel=1e-15)
    assert _line(built, "BR-01", "2000").value == pytest.approx(
        result.rwa["final_total"], rel=1e-15)


def test_sa_form_total_matches_pipeline_sa_rwa(built, result):
    assert _line(built, "BR-03", "1000").value == pytest.approx(
        result.rwa["sa"], rel=1e-9)


def test_irb_form_total_matches_pipeline_irb_rwa(built, result):
    assert _line(built, "BR-04", "1000").value == pytest.approx(
        result.rwa["irb"], rel=1e-9)


def test_lcr_form_matches_alm_engine(built, result):
    assert _line(built, "BR-08", "5000").value == pytest.approx(
        result.alm["lcr"].lcr, rel=1e-15)


def test_leverage_form_matches_engine(built, result):
    assert _line(built, "BR-07", "3000").value == pytest.approx(
        result.leverage.leverage_ratio, rel=1e-15)


def test_irrbb_outlier_flag_uses_the_method_not_the_bound_method(built, result):
    """`bool(irr.outlier)`는 바운드 메서드라 항상 참이다 — 실제 판정과 대사한다."""
    flag = _line(built, "BR-13", "3300").value
    assert flag == (1.0 if result.alm["irrbb"].outlier() else 0.0)


def test_asset_quality_form_matches_the_canonical_table(built, tables):
    aq = tables["rdm_asset_quality"]
    assert _line(built, "BR-10", "1000").value == pytest.approx(
        float(aq["balance"].sum()), rel=1e-12)
    assert _line(built, "BR-10", "1010").value == float(len(aq))


def test_reserve_requirement_is_aggregate_not_exposure_level(built, tables):
    """대손준비금은 **합계 기준**이다 — 은행업감독규정 제29조 제2항.

    이 테스트는 원래 정반대를 못박고 있었다. "총액에서 상계하면 과대충당
    익스포저가 과소충당을 가려 대손준비금이 과소산정된다"는 논리로 익스포저
    단위 합산을 지켰는데, 규정은 "대손충당금이 제1항에 따른 금액에 미달하는
    경우 그 차액"으로 합계를 대비하고 금감원 B2402-1도 그렇게 만든다.
    대손준비금은 개별 여신의 충당금이 아니라 **총 충당금을 규제 최저수준까지
    끌어올리는 이익잉여금 처분**이므로 본래 합계 개념이다.

    결함이 1차부터 8개 요청 내내 살아남은 이유가 여기 있다 — 3선이 6회 재계산해
    "일치"로 통과시켰고, 자체검증은 구조상 실패할 수 없는 항등식이 지켰고
    (지적 F-602), 이 테스트는 틀린 해석을 의도로 고정했다. 세 층이 같은 방향으로
    틀려 있었다 (지적 F-601).
    """
    from risk_lib.datamodel.materialize_detail import reserve_requirement
    rr = reserve_requirement(tables["rdm_asset_quality"])

    required = _line(built, "BR-11", "3000").value
    per_exposure = _line(built, "BR-11", "4000").value
    offset = _line(built, "BR-11", "5000").value

    assert required == pytest.approx(rr["required"], rel=1e-12)
    assert required == pytest.approx(
        max(0.0, _line(built, "BR-11", "1000").value
            - _line(built, "BR-11", "2000").value), rel=1e-12)
    # 건별 합산은 상계 효과만큼 크다. Σmax(0,xᵢ) ≥ max(0,Σxᵢ)는 정리이므로
    # 이 차이는 항상 0 이상이며, 그래서 옛 산식은 과대 방향으로만 틀렸다.
    assert per_exposure >= required - 1.0
    assert per_exposure == pytest.approx(required + offset, rel=1e-12)


def test_reserve_requirement_check_can_actually_fail():
    """BR-11 검증이 자료에 따라 실패할 수 있어야 한다 (지적 F-602).

    이전 검증은 `min(0, 건별합산 − 합계기준)`을 0과 대사했는데, 두 값의 대소가
    정리로 정해져 있어 어떤 입력에서도 정확히 0이었다. 실패 불가능성이 데이터가
    아니라 산식 구조에서 나오면 그 검증은 통제가 아니다.
    """
    from risk_lib.regulatory.forms_base import FormCheck, FormLine, _sum_check, _val
    L = [FormLine("1000", "", 0, "KRW", 128e9), FormLine("2000", "", 0, "KRW", 94e9),
         FormLine("3000", "", 0, "KRW", 80e9),   # 건별 합산을 잘못 실은 상태
         FormLine("4000", "", 0, "KRW", 80e9), FormLine("5000", "", 0, "KRW", 0.0)]
    bad = FormCheck("합계 기준", max(0.0, _val(L, "1000") - _val(L, "2000")),
                    _val(L, "3000"), 1.0)
    assert bad.status == "FAIL", "F-601 원형을 넣어도 통과하면 통제가 아니다"

    L[4] = FormLine("5000", "", 0, "KRW", 46e9)   # 분해가 어긋난 상태
    assert _sum_check("분해", L, "4000", ("3000", "5000")).status == "FAIL"


# ----- 근거·재현 --------------------------------------------------------------

def test_value_lines_carry_a_source_module_or_a_citation(built):
    for b in built:
        for ln in b.lines:
            if ln.unit == "text":
                continue
            assert ln.source_module or ln.citation, \
                f"{b.spec.form_id}/{ln.line_code} 근거 없음"


def test_digest_is_stable_and_value_sensitive(built):
    d1 = submission_digest(built)
    assert d1 == submission_digest(built)
    import copy
    mutated = copy.deepcopy(built)
    object.__setattr__(mutated[0].lines[0], "value",
                       (mutated[0].lines[0].value or 0.0) + 1.0)
    assert submission_digest(mutated) != d1


# ----- 정규 테이블 실체화 ------------------------------------------------------

def test_form_frames_satisfy_the_catalog_spec(built):
    frames = form_frames(built, "2026-06-11")
    specs = {s.name: s for s in cat.ALL_TABLES}
    for name, df in frames.items():
        bad = [v for v in validate(df, specs[name]) if v.severity == "FAIL"]
        assert bad == [], (name, bad)


def test_form_lines_reference_registered_forms(built):
    frames = form_frames(built, "2026-06-11")
    assert set(frames["reg_form_line"]["form_id"]) <= set(frames["reg_form"]["form_id"])


# ----- 엑셀 -------------------------------------------------------------------

def test_workbook_has_a_sheet_per_form_plus_control_sheets(built, tmp_path):
    p = write_workbook(built, tmp_path / "br.xlsx", asof="2026-06-11",
                       meta={"seed": 42})
    wb = load_workbook(p)
    assert wb.sheetnames[0] == "표지"
    assert wb.sheetnames[1] == "목차"
    for b in built:
        # 시트명은 서식번호다 — 감독당국이 서식번호로 찾는다.
        assert b.spec.form_no.internal_code in wb.sheetnames
    assert "검증" in wb.sheetnames and "산출근거" in wb.sheetnames


def test_workbook_values_round_trip(built, tmp_path):
    p = write_workbook(built, tmp_path / "br.xlsx", asof="2026-06-11",
                       meta={"seed": 42})
    wb = load_workbook(p, data_only=True)
    ws = wb[built[0].spec.form_no.internal_code]
    found = {}
    for row in ws.iter_rows(min_row=5, values_only=True):
        if row[0]:
            found[str(row[0])] = row[3]
    b = built[0]
    for ln in b.lines:
        if ln.unit != "text":
            assert found[ln.line_code] == pytest.approx(ln.value, rel=1e-12)


def test_workbook_cover_carries_the_digest(built, tmp_path):
    p = write_workbook(built, tmp_path / "br.xlsx", asof="2026-06-11",
                       meta={"seed": 42})
    wb = load_workbook(p)
    text = "\n".join(str(c.value) for row in wb["표지"].iter_rows()
                     for c in row if c.value)
    assert submission_digest(built) in text
    # 서식번호 전제를 표지에 남기지 않으면 배포본 매핑 시 잘못 대응된다.
    assert "서식번호" in text


def test_forms_build_when_the_month_has_no_elapsed_business_day():
    """기준일이 그 달 첫 영업일보다 앞서도 서식이 만들어진다.

    B2316(일별 트레이딩 자산)은 보고월의 경과 영업일에 잔액을 깐다. 1·2일이
    주말인 달의 기준일에서는 그 목록이 비고, 마지막 원소를 집는 코드가
    IndexError로 터지면서 **서식 290장 생성 전체가** 멈췄다. 2026-08-02(일)에
    실제로 그렇게 됐고, `report-set`·`compare`·`reproduce` CLI 검사 6건이
    한꺼번에 실패했다.

    날짜에 기대는 결함은 달력이 그 모양이 되는 날에만 나타나므로, 고정 기준일만
    쓰는 검사로는 절대 만나지 않는다. 그 날짜를 직접 박아 고정한다.
    """
    from risk_lib.data_gen import generate_portfolio
    from risk_lib.pipeline import run_pipeline
    from risk_lib.ui_studio.studio import build_studio

    p = generate_portfolio(seed=42)
    r = run_pipeline(p, asof="2026-08-02", seed=42)      # 일요일 · 월초
    s = build_studio(r, p)
    built = build_forms(r, p, s.tables)
    assert len(built) > 200

    b2316 = [b for b in built if b.spec.form_no.internal_code == "B2316"]
    assert b2316, "B2316이 생성되지 않았다"
    from risk_lib.regulatory.forms_base import _val
    days = _val(b2316[0].lines, "1000")
    assert days >= 1, "경과 영업일이 0이면 일별 잔액을 한 점도 못 깐다"


def test_overseas_rwa_denominator_carries_every_hq_component(built, result,
                                                             portfolio,
                                                             tables):
    """해외영업점 RWA 분모가 본점 RWA 구성요소를 하나도 빠뜨리지 않는다.

    빠뜨리면 **분모에서만** 사라진다 — 분자인 배분자본은 `overseas_share`로
    전액 배분되므로 해외 자본비율이 본점보다 구조적으로 높아진다. 구조화
    (집합투자증권 CRE60 · 유동화 CRE40) 4.13조가 그랬다. BF602 총자본비율이
    15.88%로 본점 10.94%를 4.94%p 웃돌았고 그 차이는 전부 누락이었다.

    라인 이름을 세지 않고 **본점 구성요소 목록**을 기준으로 본다. 파이프라인에
    갈래가 늘면(HANDOFF §7) 배분을 잊는 순간 이 검사가 먼저 깨진다.
    """
    from risk_lib.regulatory.forms_base import _val
    from risk_lib.regulatory.forms_fss_overseas_data import overseas_share
    from risk_lib.regulatory.forms import _Ctx

    rwa = result.rwa
    hq = {
        "신용리스크": float(rwa["sa"]) + float(rwa["irb"]),
        "거래상대방신용리스크": float(rwa["ccr"]),
        "시장리스크": float(rwa["market"]),
        "운영리스크": float(rwa["op"]),
        "산출하한": float(rwa["output_floor"].add_on),
        "구조화": float(rwa.get("structured_total", 0.0)),
    }
    bf602 = next(b for b in built if b.spec.form_id == "BF602")
    # 라인이 아예 없는 경우와 라인은 있는데 0인 경우를 같이 잡는다 — 후자가
    # 조용한 누락의 흔한 모양이다.
    def _allocated(name: str) -> float:
        return sum(float(ln.value or 0.0) for ln in bf602.lines
                   if name in ln.line_name and ln.unit == "KRW")

    missing = [name for name, hq_value in hq.items()
               if hq_value > 0.0 and _allocated(name) <= 0.0]
    assert not missing, (
        f"본점에 있는데 해외 분모에 없는 구성요소: {missing} — 분자만 배분되어 "
        f"해외 자본비율이 부풀려진다")

    # 구조화는 두 원장에 소재국 축이 없어 자기자본과 **같은 비중**으로 배분한다.
    # 다른 비중을 쓰면 분자·분모가 서로 다른 은행을 설명하게 된다.
    ctx = _Ctx(result, portfolio, tables)
    w = overseas_share(ctx)
    line = next(ln for ln in bf602.lines if "구조화" in ln.line_name
                and ln.unit == "KRW")
    assert line.value == pytest.approx(hq["구조화"] * w, rel=1e-12)

    # 해외 RWA 합계는 서식 자체대사(_sum_check)가 이미 보지만, 구조화를 넣고도
    # 합계에 반영하지 않는 경로가 남지 않도록 여기서도 확인한다.
    assert _val(bf602.lines, "2000") >= line.value
